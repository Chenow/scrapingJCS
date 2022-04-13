import time
import os

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from googletrans import Translator

from params import *


def initialisation_excel(mots_a_detecter=MOT_A_DETECTER):
    wb=Workbook() 
    ws=wb.active
    column=1
    classes = list(mots_a_detecter.keys())
    for i in range(1, len(mots_a_detecter) + 1):
        row = i
        ws.cell(row,column).value = classes[i - 1]
        ws.cell(row,column +1).value = 0 

    try :
        wb.save("reviews_avis"+".xlsx")  
    except : 
        wb.save("LeNomAPasMarche.xlsx")
    return wb


def initialisation_launcher(mode=MODE, amazon_url=AMAZON_URL):

    option = webdriver.ChromeOptions()
    option.add_argument("--incognito") 

    if mode == "DEV":
        option.add_argument("--start-maximized")

    if mode == "PROD": 
        option.add_argument("--headless");
    browser = webdriver.Chrome(executable_path=DRIVER_PATH, options=option)
    browser.get(amazon_url)
    button_cookies = browser.find_element_by_xpath("/html/body/div[1]/span/form/div[3]/span[1]/span/input")
    button_cookies.click()

    return browser


def get_list_avis(browser, page, url_commun=URL_COMMUN):
    browser.get(url_commun[0] + str(page) + url_commun[1] + str(page))
    try:
        list_avis = browser.find_elements_by_xpath("/html/body/div[1]/div[3]/div/div[1]/div/div[1]/div[5]/div[3]/div/div/div/div/div[4]/span/span")
    except:
        None
    return list_avis
    

def traduction(text):
    translator  = Translator()
    language = translator.detect(text).lang
    if language == "" or text == "":
        return text

    translation = translator.translate(text, src=language, dest="fr")

    return translation.text


def detection(text, mots_a_detecter=MOT_A_DETECTER):

    text = traduction(text)
    classes_du_texte = []
    classes_du_texte_sans_doublons = []
    for classe in mots_a_detecter:
        for mot in mots_a_detecter[classe]:
            if mot in text:
                classes_du_texte.append(classe)
    for classe in classes_du_texte:
        if not classes_du_texte in classes_du_texte_sans_doublons:
            classes_du_texte_sans_doublons.append(classe)

    return classes_du_texte_sans_doublons


def trier_avis(wb, list_avis, mots_a_detecter=MOT_A_DETECTER):
    
    ws=wb.active
    column = 2
    for avis in list_avis:
        classes_de_avis = detection(avis.text)
        for classe in classes_de_avis:
            row = list(mots_a_detecter.keys()).index(classe) + 1
            ws.cell(row,column).value += 1
    try :
        wb.save("reviews_avis"+".xlsx")  
    except : 
        wb.save("LeNomAPasMarche.xlsx")
    return 


